import json
import pprint
import string
import time
from pathlib import Path

from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (NoSuchElementException, StaleElementReferenceException,
										ElementNotVisibleException, TimeoutException)

class SearchPage:
	input_element = (By.ID, 'pt1:r1:0:it1::content')
	search_btn = (By.CSS_SELECTOR, 'button[class*="uenSearchButton"]')
	spinner = (By.CSS_SELECTOR, 'div[id="ShowSpinner"]')
	data_area = (By.ID, 'dataArea')

	uen = (By.CSS_SELECTOR, 'a[class*="instantInfo_link"]')
	address = (By.CSS_SELECTOR, 'a[class*="viewmap_Link"]')
	name = (By.CSS_SELECTOR, 'span.bizpara1')

class FileReader:	
	@property
	def content(self):
		path_object = Path(input("\aEnter a valid filename: "))
		if path_object.exists():
			with path_object.open() as file_handler:
				content = [ line.strip() for line in file_handler.readlines() ]
				return content if content else print("\aNo keywords in the file specified")
		else:
			print("\aYou might have to check the file name.")

class Writer:
	def __init__(self, filename):
		self.filename = filename

class XlsxWriter(Writer):
	def __init__(self, fields, filename='uen'):
		super().__init__(filename)
		self.fields = fields 
		self.letters = string.ascii_uppercase[:len(self.fields)]
		self.file_type = '.xlsx'
		self.check_filename()
		self.open_an_active_sheet()
		self.write_sheet_headers()

	def __repr__(self):
		return self.filename

	def check_filename(self):
		if self.file_type not in self.filename:
			self.filename += self.file_type
	
	def open_an_active_sheet(self):
		self.workbook = Workbook()
		self.sheet = self.workbook.active

	def close_workbook(self):
		self.workbook.save(filename=self.filename)

	def write_sheet_headers(self):
		for letter, field in zip(self.letters, self.fields):
			self.sheet[letter + str(self.sheet.max_row)].value = field

	def write_to_sheet(self, dictionary):
		try:
			max_row = str(self.sheet.max_row + 1)
			for letter, field in zip(self.letters, self.fields):
				self.sheet[letter + max_row].value = dictionary.get(field)
		finally:
			self.close_workbook()

class Baselenium:
	def __init__(self, driver_path):
		# self.driver_path = driver_path
		self.create_driver(driver_path)

	def create_driver(self, driver_path):
		'''
		creates a browser instance for selenium, 
		adds some functionalities into the browser instance
		'''
		chrome_options = Options()
		chrome_options.add_argument("start-maximized")
		chrome_options.add_argument("log-level=3")
		# the following two options are used to disable chrome browser infobar
		chrome_options.add_experimental_option("useAutomationExtension", False)
		chrome_options.add_experimental_option("excludeSwitches",["enable-automation"])

		self.driver = webdriver.Chrome(executable_path=driver_path, options=chrome_options)
		self.driver.implicitly_wait(12)

	def fetch_web_element(self, args:tuple, element=None):
		try:
			response = element.find_element(*args) if element else self.driver.find_element(*args)
		except NoSuchElementException:
			response = None
		finally:
			return response

	def fetch_web_elements(self, args:tuple, element=None):
		response = element.find_elements(*args) if element else self.driver.find_elements(*args)
		return response if response != [] else None

	def scroll_to_view(self, element):
		self.driver.execute_script("arguments[0].scrollIntoView();", element)

	def kill(self):
		self.driver.quit()

	@staticmethod
	def sift_text(element):
		if isinstance(element, webdriver.remote.webelement.WebElement):
			return element.text 

class Uen(Baselenium):
	def __init__(self, driver_path):
		super().__init__(driver_path)
		self.wait = WebDriverWait(self.driver, 45, ignored_exceptions=IGNORED_EXCEPTIONS)
		self.records = FileReader().content
		self.writer = XlsxWriter(fields)
		self.main()

	def get_page(self):
		# wait for the input element to pop up
		# if it doesn't pop up, refresh and find again
		self.driver.get("https://www.uen.gov.sg/ueninternet/faces/pages/uenSrch.jspx")

	def search_record(self, record):
		input_element = self.wait.until(EC.visibility_of_element_located(SearchPage.input_element))
		input_element.clear()
		input_element.send_keys(record)
		btn = self.fetch_web_element(SearchPage.search_btn)
		self.driver.execute_script('arguments[0].click();', btn)
		time.sleep(5)

		# wait for the spinner to go
		self.wait.until(EC.invisibility_of_element(SearchPage.spinner))
		data_area = self.wait.until(EC.visibility_of_element_located(SearchPage.data_area))
		return True if data_area.text else None

	def crawl_data(self, record):
		if (response := self.search_record(record)):
			data = dict()
			print("Looking for data")
			uen = self.wait.until(EC.presence_of_element_located(SearchPage.uen))
			if (uen := self.sift_text(uen)):
				record = uen				
			data[fields[0]] = record 

			address = self.wait.until(EC.presence_of_element_located(SearchPage.address))
			data[fields[1]] = self.sift_text(address)
			
			name = self.wait.until(EC.presence_of_all_elements_located(SearchPage.name))[6]
			data[fields[2]] = self.sift_text(name)
			pprint.pprint(data)
			self.writer.write_to_sheet(data)
		else:
			print("Can't find result ;)")

	@staticmethod
	def sleeping():
		print('sleeping . don\'t disturb :)')
		time.sleep(60)

	def main(self):
		# read in all records
		for record in self.records:
			self.get_page()
			self.sleeping()
			self.crawl_data(record)
		self.writer.close_workbook()

if __name__ == "__main__":
	IGNORED_EXCEPTIONS = (
		NoSuchElementException,
		StaleElementReferenceException,
		ElementNotVisibleException,
		TimeoutException,)

	fields = [
		'UEN',
		'ADDRESS',
		'ENTITY NAME',
	]

	driver_path = r".\chromedriver\chromedriver.exe"
	Uen(driver_path)