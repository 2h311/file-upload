import functools
import json
import logging
import pprint
import random
import string
import re
import time 
from pathlib import Path 
from itertools import count
from configparser import ConfigParser

from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (NoSuchElementException,
										StaleElementReferenceException,
										ElementNotVisibleException,
										TimeoutException)

class FileReader:	
	@property
	def content(self):
		path_object = Path(input("\aEnter a valid filename: "))
		if path_object.exists():
			with path_object.open() as file_handler:
				content = [ line.strip() for line in file_handler.readlines() ]
				return content if content else print("\aNo keywords in the file specified")
		else:
			print("\aYou might have to check the file name.")

class Writer:
	def __init__(self, filename):
		self.filename = filename

class XlsxWriter(Writer):
	def __init__(self, fields, filename='output'):
		super().__init__(filename)
		self.fields = fields 
		self.letters = string.ascii_uppercase[:len(self.fields)]
		self.file_type = '.xlsx'
		self.check_filename()
		self.open_an_active_sheet()
		self.write_sheet_headers()

	def __repr__(self):
		return self.filename

	def check_filename(self):
		if self.file_type not in self.filename:
			self.filename += self.file_type
	
	def open_an_active_sheet(self):
		self.workbook = Workbook()
		self.sheet = self.workbook.active

	def close_workbook(self):
		self.workbook.save(filename=self.filename)

	def write_sheet_headers(self):
		for letter, field in zip(self.letters, self.fields):
			self.sheet[letter + str(self.sheet.max_row)].value = field

	def write_to_sheet(self, dictionary):
		try:
			max_row = str(self.sheet.max_row + 1)
			for letter, field in zip(self.letters, self.fields):
				self.sheet[letter + max_row].value = dictionary.get(field)
		finally:
			self.close_workbook()

class Base:
	name = 'LinkedIn Navigator Locators'

class LoginPage(Base):
	iframe = By.TAG_NAME, 'iframe',
	username = By.ID, 'username',
	password = By.ID, 'password',
	signin_btn = By.CSS_SELECTOR, 'button[type="submit"]',

class SearchPage(Base):
	keywords_input = By.CSS_SELECTOR, 'input[placeholder*="Enter keywords"]',
	geography_div = By.CSS_SELECTOR, "[data-test-filter-code='GE']",
	geography_input = By.CSS_SELECTOR, 'input[placeholder="Add locations"]',
	geography_suggestion = By.CSS_SELECTOR, 'li > button',
	result_items = By.CSS_SELECTOR, '[class*="search-results__result-item"]' 

class ResultItem(Base):
	name = By.CLASS_NAME, 'result-lockup__name'
	current_workplace = By.CSS_SELECTOR, '.result-lockup__highlight-keyword',
	duration = By.CSS_SELECTOR, 'span.t-black--light', 
	location = By.CLASS_NAME, 'result-lockup__misc-item',
	previous_workplace = By.CSS_SELECTOR, '[class*="result-context__summary-list"]'
	show_more = By.CSS_SELECTOR, '[class*="result-context__past-roles-button"]'
	profile_link = By.CSS_SELECTOR, 'a'

class ResultPage(Base):
	number_of_result = By.CLASS_NAME, 'artdeco-spotlight-tab__primary-text',
	no_result = By.CLASS_NAME, 'search-results__no-results',

class ProfilePage(Base):
	name = By.CLASS_NAME, 'profile-topcard-person-entity__name',
	photo = By.CSS_SELECTOR, 'button[aria-label*=picture] img',

	entity_summary = By.CLASS_NAME, 'profile-topcard-person-entity__summary',
	summary_show_more = By.TAG_NAME, 'button',
	summary_modal_ok_btn = By.CSS_SELECTOR, '[class*="profile-topcard__summary-modal-footer"]', 
	summary_modal = By.CLASS_NAME, 'profile-topcard__summary-modal-content',

	location = By.CSS_SELECTOR, '.profile-topcard__location-data',
	connections = By.CSS_SELECTOR, '.profile-topcard__connections-data',
	contacts = By.CLASS_NAME, 'profile-topcard__contact-info-item', 
	current_workplace = By.CLASS_NAME, 'profile-topcard__current-positions',
	experience_show_more_btn = By.CSS_SELECTOR, '#profile-experience button[data-test-experience-section="expand-button"]', 
	
	positions = By.CLASS_NAME, 'profile-position',

	education_history = By.CSS_SELECTOR, 'li.profile-education',
	topcard_educations = By.CLASS_NAME, 'profile-topcard__educations',

	skills = By.ID, 'profile-skills',
	show_more_skills_btn = By.CSS_SELECTOR, 'button[data-test-skills-section="expand-button"]', 
	profile_skills = By.CLASS_NAME, 'profile-skills__pill'

	recommendations = By.CSS_SELECTOR, '.profile-recommendation-list',
	accomplishments = By.CSS_SELECTOR, '.profile-accomplishments',
	interests = By.CLASS_NAME, 'profile-interests-entity', 

	current_position = By.CLASS_NAME, 'profile-topcard__summary-position-title',
	duration = By.CLASS_NAME, 'profile-topcard__time-period-bullet',

class Baselenium:
	def __init__(self, driver_path):
		self.driver_path = driver_path
		self.create_driver()

	def set_cookies(self, filename:str, /, refresh=False):
		print("Loading cookies")
		with open(filename, 'r') as fp:
			if(contents := json.load(fp)):
				for content in contents:
					self.driver.add_cookie(content)
				if refresh:
					self.driver.refresh()
		print("Done loading cookies")

	def create_driver(self):
		'''
		creates a browser instance for selenium, 
		adds some functionalities into the browser instance
		'''
		chrome_options = Options()
		chrome_options.add_argument("start-maximized")
		chrome_options.add_argument("log-level=3")
		# the following two options are used to disable chrome browser infobar
		chrome_options.add_experimental_option("useAutomationExtension", False)
		chrome_options.add_experimental_option("excludeSwitches",["enable-automation"])
		self.driver = webdriver.Chrome(executable_path=self.driver_path, options=chrome_options)
		self.driver.implicitly_wait(30)

	def fetch_web_element(self, args:tuple, element=None):
		try:
			response = element.find_element(*args) if element else self.driver.find_element(*args)
		except NoSuchElementException:
			response = None
		finally:
			return response

	def fetch_web_elements(self, args:tuple, element=None):
		response = element.find_elements(*args) if element else self.driver.find_elements(*args)
		return response if response != [] else None

	def scroll_to_view(self, element):
		self.driver.execute_script("arguments[0].scrollIntoView();", element)

	def kill(self):
		self.driver.quit()

	@staticmethod
	def sift_text(element):
		if isinstance(element, webdriver.remote.webelement.WebElement):
			return element.text 

class ResultItemWorks:
	def name(self, dict_, li):
		name = fetch_web_element(args=ResultItem.name, element=li)
		if (text := sift_text(name)):
			dict_['Name'] = text
	
	def current_workplace(self, dict_, li):
		current_workplace = fetch_web_element(element=li, args=ResultItem.current_workplace)
		if (text := sift_text(current_workplace)):
			dict_['Current Workplace'] = text 

	def duration(self, dict_, li):
		duration = fetch_web_element(element=li, args=ResultItem.duration)
		if (text := sift_text(duration)):
			dict_['Duration'] = text
		
	def location(self, dict_, li):
		location = fetch_web_element(element=li, args=ResultItem.location)
		if (text := sift_text(location)):
			dict_['Location'] = text

	def previous(self, dict_, li):
		previous_workplace = fetch_web_element(element=li, args=ResultItem.previous_workplace)
		# check if there's a show more button
		show_more = fetch_web_element(element=li, args=ResultItem.show_more)
		logging.info('checking for a show more button in previous workplace ')
		if show_more: 	
			driver.execute_script('arguments[0].click();', show_more)
		if (text := sift_text(previous_workplace)):
			dict_['Experience/Previous Workplace'] = text
	
	def main(self, li):
		dict_ = prepopulate_dict()
		self.name(dict_, li)
		self.current_workplace(dict_, li)
		self.duration(dict_, li)
		self.location(dict_, li)
		self.previous(dict_, li)
		return dict_

def retry(function):
	'''tries to run a function after an unsuccessful attempt.'''
	@functools.wraps(function)
	def inner(*args, **kwargs):
		for _ in range(3):
			try:
				return function(*args, **kwargs)	
			except Exception as err:
				print(err)
	return inner

def config(filename='db.ini', section='navigator'):
	'''
	read in credentials
	'''
	parser = ConfigParser()
	parser.read(filename)

	# get section, defaults to navigator
	db = dict()
	if parser.has_section(section):
		params = parser.items(section)
		for param in params:
			db[param[0]] = param[1] 
	else:
		raise Exception(f"Section {section} not found in {filename}")
	return db

@retry
def login():
	# fetch the login credentials
	credentials = config()
	driver.get(base_url + '/sales/login')
	# fetch the iframe source
	iframe_src = fetch_web_element(LoginPage.iframe).get_attribute('src')
	driver.get(iframe_src)
	# enter the username
	username = fetch_web_element(LoginPage.username)
	username.send_keys(credentials.get('username'))
	# enter the password
	password = fetch_web_element(LoginPage.password)
	password.send_keys(credentials.get('password'))
	# click signin button
	fetch_web_element(LoginPage.signin_btn).click()

# this helper handles joggling between window 
switch_window = lambda handle: driver.switch_to.window(handle)

def prepopulate_dict():
	'''
	make dictionary with default values 'N/A'
	'''
	default_values = ['N/A'] * len(fields)
	return dict(list( zip(fields, default_values) ))

def trigger_extra_tab():
	'''
	trigger an extra tab to open the link profiles
	'''
	if len(driver.window_handles) == 1:
		# Open a new window
		logging.info("popping another window")
		driver.execute_script("window.open('');")
	return driver.window_handles

@retry
def interest(dict_):
	if (interests := fetch_web_elements(ProfilePage.interests)):
		scroll_to_view(interests[-1])
		interest_list = list()
		for interest in interests:
			text = interest.text.replace('\n', ': ')
			link = interest.find_element_by_tag_name('a').get_attribute('href')
			interest_list.append(text + '\n' + link)
		interests = '\n\n'.join(interest_list)
		logging.info(interests)
		dict_['Interests'] = interests

@retry
def recommendations(dict_):
	# recommendations
	if (recommendations := fetch_web_elements(ProfilePage.recommendations)):
		scroll_to_view(recommendations[-1])
		recommendations = '\n\n'.join([ recommendation.text for recommendation in recommendations ])
		logging.info(recommendations)
		dict_['Recommendations'] = recommendations

@retry
def accomplishments(dict_):
	# accomplishments
	if (accomplishments := fetch_web_elements(ProfilePage.accomplishments)):
		scroll_to_view(accomplishments[-1])
		accomplishments = '\n\n'.join([ accomplishment.text for accomplishment in accomplishments ])
		logging.info(accomplishments)
		dict_['Accomplishments'] = accomplishments

@retry
def skills(dict_):
	# # featured skills and endorsements
	profile_skills = fetch_web_element(ProfilePage.skills)
	# check if there's a show more button
	show_more = fetch_web_element(element=profile_skills, args=ProfilePage.show_more_skills_btn) 
	if show_more:
		for num in range(5):
			try:
				scroll_to_view(show_more)
				driver.execute_script('arguments[0].click();', show_more)
				break
			except Exception as err:
				print(f"An error occured: {err}") 

	skills = fetch_web_elements(element=profile_skills, args=ProfilePage.profile_skills) 
	if skills:
		scroll_to_view(skills[-1])
		skills = '\n'.join([ skill.text.replace('\n', '.') for skill in skills ])
		logging.info(skills)
		dict_['Feature Skills and Endorsement'] = skills

@retry
def experience_previous_workplace(dict_):
	# check if there's a show more button for the previous workplace 
	show_more = fetch_web_element(ProfilePage.experience_show_more_btn)
	if show_more:
		for num in range(5):
			try:
				scroll_to_view(show_more)
				driver.execute_script('arguments[0].click();', show_more)
				break
			except Exception as err:
				print(f"An error occured: {err}") 

	positions = wait.until(EC.visibility_of_all_elements_located(ProfilePage.positions))
	if positions:
		scroll_to_view(positions[-1])
		experience_previous_workplace = '\n\n'.join([ position.text for position in positions ])	
		logging.info(experience_previous_workplace)
		dict_['Experience/Previous Workplace'] = experience_previous_workplace

@retry
def education(dict_):
	# education 
	education_history = wait.until(EC.visibility_of_all_elements_located(ProfilePage.education_history))
	if education_history:
		scroll_to_view(education_history[-1])
		education_history = '\n\n'.join([ history.text for history in education_history ])
		logging.info(education_history)
		dict_['Education History'] = education_history

@retry
def current_workplace(dict_):
	# current workplace
	current_workplace = fetch_web_element(ProfilePage.current_workplace)
	if current_workplace:
		current_workplace = current_workplace.text 
		logging.info(current_workplace)
		dict_['Current Workplace'] = current_workplace

@retry
def contacts(dict_):
	contacts = fetch_web_elements(ProfilePage.contacts) 
	if contacts:
		contact_list = list()
		for contact in contacts:
			text = contact.find_element_by_tag_name('span').text
			href = contact.find_element_by_tag_name('a').get_attribute('href')
			contact_list.append(text + '\n' + href)
			contacts = '\n\n'.join(contact_list)
			logging.info(contacts)
			dict_['Contact'] = contacts

@retry
def summary(dict_):
	# fish out the summary element
	summary_element = fetch_web_element(ProfilePage.entity_summary) 
	if summary_element:
		logging.info('fishing out the summary')
		# look for the see more button
		show_more = fetch_web_element(element=summary_element, args=ProfilePage.summary_show_more)
		logging.info('checking out for a show more button to click')
		if show_more:
			driver.execute_script('arguments[0].click();', show_more)
			# check out the modal and fish out the info on it
			summary = fetch_web_element(ProfilePage.summary_modal)
			summary = sift_text(summary) 
			logging.info('closing the summary modal that popped.')
			# click the ok button to close the modal
			fetch_web_element(ProfilePage.summary_modal_ok_btn).click()
		else:
			summary = sift_text(summary_element)
		logging.info(summary)
		dict_['Summary'] = summary

@retry
def name_photo_loc_con(dict_):
	# profile page works
	name = wait.until(EC.visibility_of_element_located(ProfilePage.name)) 
	# fetch_web_element(ProfilePage.name)
	if (name := sift_text(name)):
		logging.info(name)
		dict_['Name'] = name	

	if (photo := fetch_web_element(ProfilePage.photo)):
		photo = photo.get_attribute("src") 
		# TODO: regex the photo link to make sure it's a valid link
		logging.info(photo)
		dict_['Photo'] = photo

	location = fetch_web_element(ProfilePage.location)
	if (location := sift_text(location)):
		logging.info(location)
		dict_['Location'] = location

	connections = fetch_web_element(ProfilePage.connections)
	if (connections := sift_text(connections)):
		logging.info(connections)
		dict_['Connections'] = connections

@retry
def enter_geography(geo):
	geo_div = driver.find_element_by_css_selector("[data-test-filter-code='GE']")
	for num in range(5):
		try:
			logging.info('filling the geography input')
			driver.execute_script('arguments[0].click();', geo_div)
			geo_input = geo_div.find_element_by_css_selector('input[placeholder="Add locations"]')
			geo_input.send_keys('\b'*1000)
			geo_input.send_keys(geo)		
			# click the first geography suggestion
			geo_element = fetch_web_element(element=geo_div, args=SearchPage.geography_suggestion)
			driver.execute_script('arguments[0].click();', geo_element)
			break
		except Exception as err:
			print(f"An error occured: {err}")

@retry
def enter_keyword(keyword):
	enter_keyword = wait.until(EC.visibility_of_element_located(SearchPage.keywords_input))
	enter_keyword.send_keys('\b'*1000)
	enter_keyword.send_keys(keyword)

@retry
def current_position(data_dict):
	current_position = wait.until(EC.presence_of_element_located(ProfilePage.current_position))
	if (text := sift_text(current_position)):
		data_dict['Current Position'] = text

@retry
def duration(data_dict):
	duration = wait.until(EC.visibility_of_element_located(ProfilePage.duration))
	if (text := sift_text(duration)):
		data_dict['Duration'] = text

def out_of_network(profile_link, li):
	data_dict = ResultItemWorks().main(li)
	switch_window(handles[-1])
	driver.get(profile_link)
	# fish out the education to add to the dictionary	
	if (edus := fetch_web_elements(ProfilePage.topcard_educations)):
		edus = '\n'.join([ edu.text for edu in edus ])
		data_dict['Education History'] = edus
	return data_dict

def in_network(profile_link):
	switch_window(handles[-1])
	driver.get(profile_link)
	data_dict = prepopulate_dict()
	# wait_till_page_completes()
	name_photo_loc_con(data_dict)
	duration(data_dict)
	current_position(data_dict)
	summary(data_dict)
	contacts(data_dict)
	current_workplace(data_dict)
	education(data_dict)
	experience_previous_workplace(data_dict)
	skills(data_dict)
	accomplishments(data_dict)
	recommendations(data_dict)
	interest(data_dict)
	return data_dict

def card_operations():
	# fish out the list cards
	lis = wait.until(EC.visibility_of_all_elements_located(SearchPage.result_items))
	for li in lis:
		scroll_to_view(li)
		profile_link = fetch_web_element(element=li, args=ResultItem.profile_link).get_attribute('href')
		# nap()		
		if 'OUT_OF_NETWORK' in profile_link:
			data = out_of_network(profile_link, li)
		else:
			data = in_network(profile_link)
		
		pprint.pprint(data)
		writer.write_to_sheet(data)
		switch_window(handles[0])

def traverse_pages():
	for counter in count(1):
		if counter != 1:
			url = re.sub('page=\d+', f'page={counter}', driver.current_url)
			print(url)
			driver.get(url)
	
		element = fetch_web_element(ResultPage.no_result)
		if not element:
			card_operations()
		else:
			print("No Result Found...")
			break

def nap():
	secs = random.randrange(3, 12)
	time.sleep(secs)

IGNORED_EXCEPTIONS = (
	NoSuchElementException,
	StaleElementReferenceException,
	ElementNotVisibleException,
	TimeoutException,)

fields =[
	'Name',
	'Photo',
	'Connections',
	'Contact',
	'Summary',
	'Current Workplace',
	'Experience/Previous Workplace',
	'Education History',
	'Feature Skills and Endorsement',
	'Recommendations',
	'Accomplishments',
	'Interests',
	'Current Position',
	'Duration',
	'Location',
]

# if __name__ == "__main__":
# driver_path = r"C:\Users\DELL\jobs\chromedriver\chromedriver.exe"
driver_path = "./chromedriver/chromedriver.exe"
base_url = 'https://www.linkedin.com'
logging.basicConfig(format="## %(message)s", level=logging.INFO)

# logging.disable(logging.INFO)
bs = Baselenium(driver_path)
driver = bs.driver
fetch_web_element = bs.fetch_web_element
fetch_web_elements = bs.fetch_web_elements
sift_text = bs.sift_text
scroll_to_view = bs.scroll_to_view
wait = WebDriverWait(driver, 60, ignored_exceptions=IGNORED_EXCEPTIONS)
writer = XlsxWriter(fields)

handles = trigger_extra_tab()
switch_window(handles[0])

login()
# uncomment this to use coookies instead of the login procedure
# driver.get(base_url)
# bs.set_cookies('cookies.json')

if (keywords := FileReader().content):
	for key in keywords:
		keyword, geo = key.split(',')
		driver.get(base_url + '/sales/search/people')
		enter_keyword(keyword)
		enter_geography(geo)
		traverse_pages()
